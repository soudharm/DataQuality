import argparse
import configparser as cp
import csv
from openpyxl import load_workbook
# import pyarrow.orc as orc
import pathlib
import pandas as pd

#from DQF.scripts.upload_config_file import azureupload
from scripts.cloud_storages import establish_db_conn, connect_aws_storage, azure_datalake_storage, read_file_adls
from scripts.rule_check import *
from scripts.statistics_data_writing import *
from scripts.email_notification import *
from scripts.upload_config_file import *

timeString = str(datetime.now().strftime("%Y-%m-%d_%H-%M-%S"))
df_dict = {}

def read_args():
    parser = argparse.ArgumentParser()
    parser.add_argument("--project", required=True)
    parser.add_argument("--source_type", choices=["file", "table"], required=False)
    parser.add_argument("--source_name", nargs='+', required=False)
    parser.add_argument("--config_location", choices=["s3", "adls"], required=False)
    args = parser.parse_args()
    return args


def folder_paths():
    args = read_args()
    project = args.project
    config_location = args.config_location
    conf_obj = cp.ConfigParser()
    main_project_path = os.getcwd()
    project_path = os.path.join(main_project_path, "DQF\Projects", project)
    #print(project_path)
    full_path = os.path.join(project_path, "config") + r"\framework_config.ini"
    conf_obj.read(full_path)
    output = conf_obj["File_Configuration"]["output_folder_path"]
    today = datetime.now()
    year = today.strftime("%Y")
    month = today.strftime("%m")
    day = today.strftime("%d")
    output_folder_path = project_path + output + year + "-" + month + "-" + day
    if not os.path.exists(output_folder_path):
        os.makedirs(output_folder_path)
    output_file_path = str(output_folder_path) + r"\ValidationResult_{}.xlsx".format(timeString)
    # detailed_output_file = str(output_folder_path) + r"\Detailed_result_{}.csv".format(timeString)
    if config_location == 's3':
        rule_config_file, admin_file = connect_aws_storage(conf_obj)
    elif config_location == 'adls':
        rule_config_file, admin_file = azure_datalake_storage(conf_obj)
    else:
        config_path = conf_obj["File_Configuration"]["config_location"]
        config_file = project_path + config_path
        rule_config_file = pd.read_excel(config_file, sheet_name=["DQ_RULE_CONFIG"])
        #admin_file = pd.read_excel(config_file, sheet_name=["DQ_RULE_ADMIN"])
        admin_file = pd.read_excel(config_file, sheet_name=["DQ_RULE_CONFIG"])

    return conf_obj, output_folder_path, output_file_path, project_path, rule_config_file, admin_file


def validating_config_file(conf_obj, rule_config_file, admin_file, config_count, conf_file_path, output_file_path,  log_file_path):
    rule_config_file_merged = []
    args = read_args()
    source_type = args.source_type
    source_name = args.source_name
    rule_config_file_df = rule_config_file["DQ_RULE_CONFIG"]
    if source_name:
        for source in source_name:
            if source_type:
                if source_type == "table":
                    rule_config_file = \
                        rule_config_file_df[(rule_config_file_df['ACTIVE_FLAG'] == 'Y') & (
                                rule_config_file_df['validation_table'] == source)]
                    rule_config_file_merged.append(rule_config_file)
                else:
                    rule_config_file = \
                        rule_config_file_df[(rule_config_file_df['ACTIVE_FLAG'] == 'Y') & (
                                rule_config_file_df['source_file'] == source)]
                    rule_config_file_merged.append(rule_config_file)
            else:
                rule_config_file_table = rule_config_file_df[(rule_config_file_df['ACTIVE_FLAG'] == 'Y') & (
                            rule_config_file_df['validation_table'] == source)]
                rule_config_file_merged.append(rule_config_file_table)
                rule_config_file_file = rule_config_file_df[(rule_config_file_df['ACTIVE_FLAG'] == 'Y') & (
                            rule_config_file_df['source_file'] == source)]
                rule_config_file_merged.append(rule_config_file_file)
    elif source_type and not source_name:
        rule_config_file_merged = [rule_config_file_df[(rule_config_file_df['ACTIVE_FLAG'] == 'Y') & (
                rule_config_file_df['rule_action_type'] == source_type) & (rule_config_file_df['config_id'] > int(config_count))]]
    else:
        rule_config_file_merged = [rule_config_file_df[(rule_config_file_df['ACTIVE_FLAG'] == 'Y') & (rule_config_file_df['config_id'] > int(config_count))]]
    admin_config_df = pd.concat(rule_config_file_merged)
    ##admin_config_df = pd.merge(admin_file["DQ_RULE_ADMIN"], rule_config_file_merged, suffixes=('_', ''), on='rule_id',
                               #how='inner')
    if not admin_config_df.empty:
        sorted_df = admin_config_df.reindex(
            admin_config_df.config_id.astype(float).astype(int).sort_values().index).reset_index(drop=True)
        validate_config_rows(sorted_df, conf_obj)
        if config_count == str(0):
            if conf_file_path != str(0):
                pass
            else:
                write_output_data(sorted_df, output_file_path)
                with open(log_file_path, 'w') as log_file:
                    log_file.write(str(0) + "," + output_file_path)
    else:
        raise ValueError("Empty Dataset Found")
    return sorted_df


def read_source_file(source_file_path):
    file_extension = pathlib.Path(source_file_path).suffix
    dataset = ""
    if file_extension == '.csv':
        dataset = pd.read_csv(source_file_path)
    elif file_extension in ['.xls', '.xlsx', '.xlsm', '.xlsb', '.odf', '.ods', '.odt']:
        dataset = pd.read_excel(source_file_path)
    elif file_extension == ".json":
        dataset = pd.read_json(source_file_path)
    elif file_extension == ".xml":
        dataset = pd.read_xml(source_file_path)
    elif file_extension == ".parquet":
        dataset = pd.read_parquet(source_file_path, engine="pyarrow")
    # elif file_extension == ".orc":
    #     with open(source_file_path) as file:
    #         data = orc.ORCFile(source_file_path)
    #         dataset = data.read().to_pandas()
    return dataset


def write_output_data(output_message, out_file_path):
    """
    This is function is used to generating csv file for output
    """
    conf_obj, output_folder_path, output_file_path, project_path, rule_config_file, admin_file = folder_paths()
    sheet_name = "validation_summary"
    file_exists = os.path.exists(out_file_path)
    header = ['Config ID', 'Rule Name', 'File Path', 'Table', 'Validation Column', 'Msg/Comment',
              'No. Of Rows Impacted']
    if not file_exists:
        with pd.ExcelWriter(out_file_path) as writer:
            df = pd.DataFrame(columns=header, index=None)
            df.to_excel(writer, sheet_name, index=False, header=True)
        write_statistics_result(output_message, out_file_path, conf_obj)
    else:
        workbook = load_workbook(out_file_path)
        writer = pd.ExcelWriter(out_file_path, engine='openpyxl')
        writer.book = workbook
        writer.sheets = dict((ws.title, ws) for ws in workbook.worksheets)
        output_row = output_message[0:len(output_message) - 1]
        data = [{'Config ID': output_row[0], 'Rule Name': output_row[2], 'File Path': output_row[3], 'Table': output_row[4],
                 'Validation Column': output_row[5],
                 'Msg/Comment': output_row[6], 'No. Of Rows Impacted': output_row[7]}]
        df1 = pd.DataFrame.from_dict(data)
        df1.to_excel(writer, sheet_name, startrow=writer.sheets[sheet_name].max_row, index=False,
                     header=False)
        writer.save()


def write_detailed_output(output_message, out_file_path):
    """
     To generate detailed output of data
    """
    detailed_list = output_message[len(output_message) - 2]
    if isinstance(detailed_list, pd.DataFrame) or detailed_list:
        dq_check_action = output_message[9]
        config_id = output_message[0]
        rule_name = output_message[2]
        validation_column = output_message[5]
        sheet_name = str(rule_name) + "_" + str(validation_column) + "_" + str(config_id)
        workbook = load_workbook(out_file_path)
        writer = pd.ExcelWriter(out_file_path, engine='openpyxl')
        writer.book = workbook
        writer.sheets = dict((ws.title, ws) for ws in workbook.worksheets)
        df1 = pd.DataFrame.from_dict(detailed_list)
        df1.to_excel(writer, sheet_name, index=False, header=True)
        writer.save()
        if dq_check_action:
            raise ValueError(output_message[6])


def write_statistics_result(admin_config, output_file_path, conf_obj):
    dataset = ""
    sheet_name = ""
    main_project_path = os.getcwd()
    args = read_args()
    project = args.project
    project_path = os.path.join(main_project_path, "Projects", project)
    workbook = load_workbook(output_file_path)
    writer = pd.ExcelWriter(output_file_path, engine='openpyxl')
    writer.book = workbook
    df = admin_config.drop_duplicates(['source_file', 'validation_table', 'Platform'])
    df.reset_index(inplace=True)
    for i in range(len(df)):
        if df['rule_action_type'][i] == "file":
            source_file = df["source_file"][i]
            if df['rule_name'][i] not in ["file_count_check", "dataset_equality_check", "dataset_length_check", "dataset_content_check"] and "*" not in source_file:
                if admin_config['Source_file_location'][i] == "adls":
                    dataset = read_file_adls(conf_obj, source_file)
                    sheet_name = "stats_" + source_file.split("/")[-1]
                else:
                    source_file_path = project_path + conf_obj["File_Configuration"]["input_folder_path"] + source_file
                    dataset = read_source_file(source_file_path)
                    sheet_name = "stats_" + source_file
            elif "*" in source_file:
                source_file = source_file.replace('*', '')
                for (root, dirs, file) in os.walk(source_file):
                    for f in file:
                        if admin_config['Source_file_location'][i] == "adls":
                            dataset = read_file_adls(conf_obj, source_file)
                            sheet_name = "stats_" + source_file.split("/")[-1]
                        else:
                            dataset = read_source_file(root + "\\" + f)
                            sheet_name = "stats_" + f
                        stats_df = statistics_of_df(dataset)
                        writer.sheets = dict((ws.title, ws) for ws in workbook.worksheets)
                        stats_df.to_excel(writer, sheet_name, index=False, header=True)
                        writer.save()
        elif df['rule_action_type'][i] == "table":
            table = df["validation_table"][i]
            platform_name = df['Platform'][i]
            if "," in table:
                table_names = [a.strip() for a in table.split(',')]
                for t in table_names:
                    dataset = establish_db_conn(conf_obj, t, platform_name)
                    sheet_name = "stats_" + t + "_" + platform_name
                    stats_df = statistics_of_df(dataset)
                    writer.sheets = dict((ws.title, ws) for ws in workbook.worksheets)
                    stats_df.to_excel(writer, sheet_name, index=False, header=True)
                    writer.save()
            else:
                dataset = establish_db_conn(conf_obj, table, platform_name)
                sheet_name = "stats_" + table + "_" + platform_name
        if isinstance(dataset, pd.DataFrame):
            stats_df = statistics_of_df(dataset)
            print(stats_df, 222222222222222222222222222)
            writer.sheets = dict((ws.title, ws) for ws in workbook.worksheets)
            stats_df.to_excel(writer, sheet_name, index=False, header=True)
            writer.save()

def check_rules(admin_config, i, rule_id, validation_column, execution_parameter, out_file_path):
     source_file = ""
     table = ""
     source_path = ""
     dataset = ""
     conf_obj, output_folder_path, output_file_path, project_path, rule_config_file, admin_file = folder_paths()
     if admin_config['rule_action_type'][i] == "file":
          source_file = admin_config["source_file"][i]
          if admin_config['rule_name'][i] in ["file_count_check", "dataset_equality_check", "dataset_length_check", "dataset_content_check"]:
              call_checks(admin_config, i, rule_id, validation_column, execution_parameter, out_file_path, " ",
                          source_file, source_path, table, conf_obj)
          elif "*" in source_file:
               source_file = source_file.replace('*', '')
               for (root, dirs, file) in os.walk(source_file):
                   for f in file:
                       if admin_config['Source_file_location'][i] == "local":
                           source_file = root + "\\" + f
                           dataset = read_source_file(source_file)
                       elif admin_config['Source_file_location'][i] == "adls":
                           dataset = read_file_adls(conf_obj, source_file)
                       call_checks(admin_config, i, rule_id, validation_column, execution_parameter, out_file_path, dataset, source_file,source_path, table, conf_obj)
          else:
              if admin_config['Source_file_location'][i] == "local":
                  source_path = project_path + conf_obj["File_Configuration"]["input_folder_path"]
                  source_file_path = source_path + str(source_file)
                  dataset = read_source_file(source_file_path)
              elif admin_config['Source_file_location'][i] == "adls":
                  dataset = read_file_adls(conf_obj, source_file)
              call_checks(admin_config, i, rule_id, validation_column, execution_parameter, out_file_path, dataset, source_file,source_path, table, conf_obj)
     elif admin_config['rule_action_type'][i] == "table":
          table = admin_config["validation_table"][i]
          platform_name = admin_config['Platform'][i]
          dataset = establish_db_conn(conf_obj, table, platform_name)
          call_checks(admin_config, i, rule_id, validation_column, execution_parameter, out_file_path, dataset, source_file,source_path, table, conf_obj)


def call_checks(admin_config, i, rule_id, validation_column, execution_parameter, out_file_path, dataset, source_file,source_path, table, conf_obj):
    detailed_output_check = admin_config["detailed_output"][i]

    # Column Nullability Check
    if admin_config["rule_name"][i] == 'not_null':
        output_message = column_nullability_check(admin_config["config_id"][i], rule_id,
                                                dataset,
                                                validation_column, source_file, table,
                                                execution_parameter, admin_config["DQ_CHECK_STAGE"][i])
        write_output_data(output_message, out_file_path)
        if detailed_output_check == "Y":
           write_detailed_output(output_message, out_file_path)


    # Checking Duplicate
    elif admin_config["rule_name"][i] == 'unique':

        output_message = duplicate_check(admin_config["config_id"][i], rule_id, dataset,
                                       validation_column,
                                       source_file, table, execution_parameter, admin_config["DQ_CHECK_STAGE"][i])
        write_output_data(output_message, out_file_path)
        if detailed_output_check == "Y":
           write_detailed_output(output_message, out_file_path)

    # Check integer datatype
    elif admin_config["rule_name"][i] == 'int_check':
        output_message = integer_datatype_check(admin_config["config_id"][i], rule_id, dataset,
                                              validation_column,
                                              source_file, table, execution_parameter,
                                              admin_config["DQ_CHECK_STAGE"][i])
        write_output_data(output_message, out_file_path)
        if detailed_output_check == "Y":
           write_detailed_output(output_message, out_file_path)

    # Check char datatype
    elif admin_config["rule_name"][i] == 'char_check':
        output_message = char_datatype_check(admin_config["config_id"][i], rule_id, dataset,
                                           validation_column,
                                           source_file, table, execution_parameter,
                                           admin_config["DQ_CHECK_STAGE"][i])
        write_output_data(output_message, out_file_path)
        if detailed_output_check == "Y":
           write_detailed_output(output_message, out_file_path)

    # Check varchar datatype
    elif admin_config["rule_name"][i] == 'varchar_check':
        output_message = varchar_datatype_check(admin_config["config_id"][i], rule_id, dataset,
                                              validation_column,
                                              source_file, table, execution_parameter,
                                              admin_config["DQ_CHECK_STAGE"][i])
        write_output_data(output_message, out_file_path)
        if detailed_output_check == "Y":
           write_detailed_output(output_message, out_file_path)

    # Check boolean datatype
    elif admin_config["rule_name"][i] == 'boolean_check':
        output_message = boolean_datatype_check(admin_config["config_id"][i], rule_id, dataset,
                                              validation_column,
                                              source_file, table, execution_parameter,
                                              admin_config["DQ_CHECK_STAGE"][i])
        write_output_data(output_message, out_file_path)
        if detailed_output_check == "Y":
           write_detailed_output(output_message, out_file_path)

    # Check values in specified decimal points
    elif admin_config["rule_name"][i] == 'decimal_check':
        output_message = decimal_check(admin_config["config_id"][i], rule_id, dataset,
                                     validation_column,
                                     source_file, table, execution_parameter, admin_config["DQ_CHECK_STAGE"][i])
        write_output_data(output_message, out_file_path)
        if detailed_output_check == "Y":
           write_detailed_output(output_message, out_file_path)

    # Check values in specified list
    elif admin_config["rule_name"][i] == 'lst_values_check':
        output_message = specified_list_check(admin_config["config_id"][i], rule_id, dataset,
                                            validation_column,
                                            source_file, table, execution_parameter,
                                            admin_config["DQ_CHECK_STAGE"][i])
        write_output_data(output_message, out_file_path)
        if detailed_output_check == "Y":
           write_detailed_output(output_message, out_file_path)

    # Check values in specified dateformat type
    elif admin_config["rule_name"][i] == 'date_format_check':
        output_message = dateformat_check(admin_config["config_id"][i], rule_id, dataset,
                                        validation_column,
                                        source_file, table, execution_parameter, admin_config["DQ_CHECK_STAGE"][i])
        write_output_data(output_message, out_file_path)
        if detailed_output_check == "Y":
           write_detailed_output(output_message, out_file_path)

    # Check values in specified time type
    elif admin_config["rule_name"][i] == 'timestamp_check':
        output_message = time_check(admin_config["config_id"][i], rule_id, dataset,
                                  validation_column,
                                  source_file, table, execution_parameter, admin_config["DQ_CHECK_STAGE"][i])
        write_output_data(output_message, out_file_path)
        if detailed_output_check == "Y":
           write_detailed_output(output_message, out_file_path)

    # DataSet size check
    elif admin_config["rule_name"][i] == 'file_size':
        source_file_path = source_path + str(source_file)
        output_message = dataset_size_check(admin_config["config_id"][i], rule_id, dataset,
                                          source_file_path, source_file, table,
                                          execution_parameter, admin_config["DQ_CHECK_STAGE"][i])
        write_output_data(output_message, out_file_path)

    # dataset file extension check
    elif admin_config["rule_name"][i] == 'file_extension':
        output_message = file_extension_check(admin_config["config_id"][i], rule_id, dataset,
                                            source_file, table,
                                            execution_parameter, admin_config["DQ_CHECK_STAGE"][i])
        write_output_data(output_message, out_file_path)

    # count dataset column
    elif admin_config["rule_name"][i] == 'file_col_count_check':
        output_message = dataset_column_count(admin_config["config_id"][i], rule_id, dataset,
                                            source_file, table,
                                            execution_parameter, admin_config["DQ_CHECK_STAGE"][i])
        write_output_data(output_message, out_file_path)

    elif admin_config["rule_name"][i] == 'header_pattern_check':
        output_message = header_pattern_check(admin_config["config_id"][i], rule_id, dataset,
                                            validation_column,
                                            source_file, table, execution_parameter,
                                            admin_config["DQ_CHECK_STAGE"][i])
        write_output_data(output_message, out_file_path)

    elif admin_config["rule_name"][i] == 'pattern_check':
        output_message = pattern_check(admin_config["config_id"][i], rule_id, dataset,
                                     validation_column,
                                     source_file, table, execution_parameter, admin_config["DQ_CHECK_STAGE"][i])
        write_output_data(output_message, out_file_path)

    elif admin_config["rule_name"][i] == 'file_availability_check':
        output_message = file_availability_check(admin_config["config_id"][i], rule_id, dataset,
                                               validation_column,
                                               source_path, table, execution_parameter,
                                               admin_config["DQ_CHECK_STAGE"][i])
        write_output_data(output_message, out_file_path)

    elif admin_config["rule_name"][i] == 'file_folder_availability_check':
        output_message = file_folder_availability_check(admin_config["config_id"][i], rule_id, dataset,
                                                      validation_column,
                                                      source_path, table, execution_parameter,
                                                      admin_config["DQ_CHECK_STAGE"][i])
        write_output_data(output_message, out_file_path)

    elif admin_config["rule_name"][i] == 'relationship':
        platform_name = admin_config["Platform"][i]
        output_message = relationship(admin_config["config_id"][i], rule_id, validation_column, source_file, table, platform_name, admin_config["DQ_CHECK_STAGE"][i],
                  conf_obj)
        write_output_data(output_message, out_file_path)
        if detailed_output_check == "Y":
           write_detailed_output(output_message, out_file_path)

    elif admin_config["rule_name"][i] == 'file_count_check':
        output_message = file_count(admin_config["config_id"][i], rule_id, source_path, source_file, table, execution_parameter, admin_config["DQ_CHECK_STAGE"][i])
        write_output_data(output_message, out_file_path)
        if detailed_output_check == "Y":
           write_detailed_output(output_message, out_file_path)

    # Check values in specified length
    elif admin_config["rule_name"][i] == 'column_length_check':
        output_message = column_length_check(admin_config["config_id"][i], rule_id, dataset,
                                       validation_column,
                                       source_file, table, execution_parameter, admin_config["DQ_CHECK_STAGE"][i])
        write_output_data(output_message, out_file_path)
        if detailed_output_check == "Y":
            write_detailed_output(output_message, out_file_path)

    # file rows check between source and destination
    elif admin_config["rule_name"][i] == 'dataset_equality_check':
        output_message = dataset_equality_check(admin_config["config_id"][i], rule_id, source_path, source_file, table, execution_parameter, admin_config["DQ_CHECK_STAGE"][i])
        write_output_data(output_message, out_file_path)
        if detailed_output_check == "Y":
           write_detailed_output(output_message, out_file_path)

    # file length check between source and destination
    elif admin_config["rule_name"][i] == 'dataset_length_check':
        output_message = dataset_length_check(admin_config["config_id"][i], rule_id, source_path, source_file, table, execution_parameter, admin_config["DQ_CHECK_STAGE"][i])
        write_output_data(output_message, out_file_path)
        if detailed_output_check == "Y":
           write_detailed_output(output_message, out_file_path)

    # file content check between source and destination
    elif admin_config["rule_name"][i] == 'dataset_content_check':
        output_message = dataset_content_check(admin_config["config_id"][i], rule_id, source_path, source_file, table, execution_parameter, admin_config["DQ_CHECK_STAGE"][i])
        write_output_data(output_message, out_file_path)
        if detailed_output_check == "Y":
           write_detailed_output(output_message, out_file_path)


def validate_config_rows(rule_config_file, conf_obj):
    rule_file = rule_config_file
    for i in range(len(rule_file['rule_name'])):
        if (rule_file['rule_action_type'][i] == 'table') & (type(rule_file['validation_table'][i]) is not str) & (rule_file['ACTIVE_FLAG'][i] == 'Y'):
            raise ValueError("Missing table name where config_id => " + str(rule_file['config_id'][i]))
        if (rule_file['rule_action_type'][i] == 'file') & (type(rule_file['source_file'][i]) is not str) &\
                    (rule_file['rule_name'][i] not in ['file_availability_check']) & (rule_file['ACTIVE_FLAG'][i] == 'Y'):
            raise ValueError("Missing source file path where config_id => " + str(rule_file['config_id'][i]))
        if (rule_file['rule_action_type'][i] == 'table') & (type(rule_file['Platform'][i]) is not str) & (rule_file['ACTIVE_FLAG'][i] == 'Y'):
            raise ValueError("Missing Platform where config_id => " + str(rule_file['config_id'][i]))
        if (rule_file['rule_name'][i] in ('date_format_check', 'lst_values_check')) & (
                type(rule_file['Exec_Parameter'][i]) is not str) & (rule_file['ACTIVE_FLAG'][i] == 'Y'):
            raise ValueError("Missing Exec_Parameter where config_id => " + str(rule_file['config_id'][i]))
        if (type(rule_file['rule_name'][i]) is not str) or (type(rule_file['rule_id'][i]) is not str) or \
                (type(rule_file['rule_action_type'][i]) is not str):
            raise ValueError("Missing values found in Config File where config_id => " + str(rule_file['config_id'][i]))
        # if rule_file['rule_action_type'][i] == 'table':
        #     if type(rule_file['validation_table'][i]) is not str:
        #         raise ValueError("Missing table name where config_id => " + str(rule_file['config_id'][i]))
        #     else:
        #         output = establish_db_conn(conf_obj, rule_file['validation_table'][i], rule_file['Platform'][i])
        #         if output is False:
        #             raise ValueError(rule_file['validation_table'][i] + " Not Found where config_id => " + str(rule_file['config_id'][i]))

def read_from_config():
    """
    This Function read the Config sheet and Rule Sheet
    """
    conf_obj, output_folder_path, output_file_path, project_path, rule_config_file, admin_file = folder_paths()
    log_file_path = project_path + "\\" + "config\log_config.txt"
    with open(log_file_path, 'r') as log_file:
        row = log_file.readline()
        config_log_data = [a.strip() for a in row.split(',')]
        config_count = config_log_data[0]
        conf_file_path = config_log_data[1]
        if conf_file_path != output_file_path and conf_file_path != str(0):
            output_file_path = conf_file_path
        else:
            output_file_path = output_file_path
    admin_config = validating_config_file(conf_obj, rule_config_file, admin_file, config_count,conf_file_path, output_file_path, log_file_path)
    # df = admin_config.drop_duplicates(['source_file', 'validation_table', 'Platform'])
    # df.reset_index(inplace=True)
    for i in range(len(admin_config["rule_id"])):
        execution_parameter = admin_config['Exec_Parameter'][i]
        validation_column = admin_config["validation_column"][i]
        rule_id = admin_config["rule_id"][i]
        rule_name = admin_config["rule_name"][i]
        config_id = admin_config["config_id"][i]
        if type(execution_parameter) is str:
            if admin_config['rule_action_type'][i] == "file":
                source_file = admin_config["source_file"][i]
                if type(source_file) is not str:
                    print("Error: source file path not provided, ConfigID = > ", config_id)
                else:
                    check_rules(admin_config, i, rule_id, validation_column, execution_parameter, output_file_path)
            elif admin_config['rule_action_type'][i] == "table":
                table = admin_config["validation_table"][i]
                if type(table) is not str:
                    print("Error: Table not provided, ConfigID = > ", config_id)
                else:
                    check_rules(admin_config, i, rule_id, validation_column, execution_parameter, output_file_path)
            else:
                check_rules(admin_config, i, rule_id, validation_column, execution_parameter, output_file_path)
        else:
            check_rules(admin_config, i, rule_id, validation_column, execution_parameter, output_file_path)
        print(f"Successfully Executed Config ID => {config_id}, Rule ID => {rule_id}, Rule Name => {rule_name}")
        with open(log_file_path, 'w') as log_file:
            log_file.write(str(config_id) + "," + output_file_path)
    with open(log_file_path, 'w') as log_file:
        log_file.write(str(0) + "," + str(0))
    send_email_notification(output_file_path, project_path)
    azureupload(output_file_path,project_path)

def updateconfig():
    connection_string = "DefaultEndpointsProtocol=https;AccountName=diadifstorage001;AccountKey=kmNicxR5p69wOAHgIj8zqgA5Ky3Ej0xDU9haRELi0zG9aZv45VlHUoqPU9Wi7V6IOu8zDT1UgZ97+AStwFUxWw==;EndpointSuffix=core.windows.net"
    file_path = "fca/config/Config_Rule.xlsx"
    try:
        rule_config_file = pd.read_excel(
            f"abfs://{file_path}",
            storage_options={
                "connection_string": connection_string
            }, na_values = "Missing",sheet_name="DQ_RULE_CONFIG")
        now = datetime.now()
        date_today = now.strftime("%Y%m%d")

        for i in range(len(rule_config_file)):
            if (rule_config_file['Source_file_location'][i] == 'adls') & (rule_config_file['source_file'].str.contains('part*').any()):
                start_e = rule_config_file['source_file'][i].rindex('/',1,-6)
                end_e = rule_config_file['source_file'][i].rindex('.')
                rule_config_file.source_file[i] = rule_config_file['source_file'][i].replace(
                    rule_config_file['source_file'][i][start_e + 1:end_e], date_today)
                #print(rule_config_file.source_file[i])
            elif (rule_config_file['Source_file_location'][i] == 'adls') & (rule_config_file['source_file'][i] != None):
                start = rule_config_file['source_file'][i].rindex('/')
                end = rule_config_file['source_file'][i].rindex('.')
                rule_config_file.source_file[i]=rule_config_file['source_file'][i].replace(rule_config_file['source_file'][i][start + 1:end], date_today)
                print(rule_config_file.source_file[i])


        rule_config_file.to_excel(
            f"abfs://{file_path}",
            storage_options={
                "connection_string": connection_string
            },sheet_name="DQ_RULE_CONFIG",index=False)

        ##print(rule_config_file)

    except Exception as e:
        print(e)



# Main function
if __name__ == '__main__':
    import warnings
    warnings.filterwarnings("ignore")
    #updateconfig()
    read_from_config()
    #print(os.getcwd())
    # args = read_args()
    # project = args.project
    # config_location = args.config_location
    # conf_obj = cp.ConfigParser()
    # # main_project_path = os.getcwd()
    # project_path = os.path.join(os.getcwd(), "DQF/Projects", "azure")
    # # print(project_path)
    # full_path = os.path.join(project_path, "config") + r"\framework_config.ini"
    # conf_obj.read(full_path)
    # output = conf_obj["File_Configuration"]["output_folder_path"]
    # today = datetime.now()
    # year = today.strftime("%Y")
    # month = today.strftime("%m")
    # day = today.strftime("%d")
    # output_folder_path = project_path + r"\\" + output + r"\\" + year + "-" + month + "-" + day
    # if not os.path.exists(output_folder_path):
    #     os.makedirs(output_folder_path)
    # output_file_path = str(output_folder_path) + r"\ValidationResult_{}.xlsx".format(timeString)
    # print(output_file_path)

